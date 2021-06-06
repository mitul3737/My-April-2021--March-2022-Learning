import openpyxl
wb=openpyxl.load_workbook('produceSales.xlsx')
sheet=wb['Sheet'] # using the sheet named'Sheet'

#updating values as we want in a dictionary named PRICE_UPDATE
PRICE_UPDATE={'Garlic':3.07,
              'Celery':1.19,
              'Lemon':1.27}


for rowNum in range(2,sheet.max_row):
    produceName=sheet.cell(row=rowNum,column=1).value #loop starts rom row 2 as there are all the values
    if produceName in PRICE_UPDATE:  #searching the keys that are in the dictionary PRICE_UPDATE
        sheet.cell(row=rowNum,column=2).value=PRICE_UPDATE[produceName]  #PRICE_UPDATE[produceName] will give the key's value assigned to the desired cell



wb.save('updatedProductSales.xlsx') # creating a new file named updatedProductSales.xlsx and saving it
