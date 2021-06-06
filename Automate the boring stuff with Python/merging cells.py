import openpyxl
wb=openpyxl.Workbook()
sheet=wb.active
sheet.merge_cells('A1:D3')
sheet['A1']='Twelve cells merged together'
sheet.merge_cells('C5:D5')
sheet['C5']="Two merged cells."
print(wb.save('merged.xlsx'))